from openpyxl import load_workbook
import matplotlib.pyplot as plt

wb = load_workbook('../code/Financial Sample.xlsx')


def init_spreadsheet():
    return wb['Finances 2017']


# I want the percentage of units sold in each country
def percentage_of_units_sold(ws1):
    maxrow = ws1.max_row
    countries = {}
    for row in range(2, maxrow):
        cell = "".join(['E', str(row)])
        country = "".join(['B', str(row)])
        if ws1[country].value in countries.keys():
            countries[ws1[country].value] += ws1[cell].value
        else:
            countries[ws1[country].value] = ws1[cell].value
    names = countries.keys()
    countries['USA'] = countries.get('United States of America')
    del countries['United States of America']
    total_units = sum(countries.values())
    values = [value/total_units*100 for value in countries.values()]
    plt.figure(figsize=(25, 9))
    plt.subplot(131)
    plt.bar(names, values)
    plt.savefig("porcentage_units.png")


if __name__ == '__main__':
    ws1 = init_spreadsheet()
    percentage_of_units_sold(ws1)
    wb.save('../code/Financial Sample.xlsx')
